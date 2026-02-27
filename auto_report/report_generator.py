"""
基金申赎月度报告生成器
- 读取原始数据 → 月度汇总表
- 格式化Excel
- 插入AI摘要文本
- 趋势图表
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

# ── 常量 ──────────────────────────────────────────────
MAJOR_CATEGORIES = ['1权益（风格标签）', '2纯债（券种偏好）', '3固收+', '5 QDII基金', '8其他-香港互认']

HEADER_FILL = PatternFill(start_color='5B9BD5', fill_type='solid')
TITLE_FILL = PatternFill(start_color='4472C4', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='E7E6E6', fill_type='solid')
SUMMARY_HEADER_FILL = PatternFill(start_color='4472C4', fill_type='solid')
# SUMMARY_BG_FILL = PatternFill(start_color='F2F7FB', fill_type='solid')
SUMMARY_BG_FILL = PatternFill(fill_type=None)

WHITE_BOLD = Font(size=11, bold=True, color='FFFFFF', name='Arial')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial')
TOTAL_FONT = Font(bold=True, name='Arial')
NORMAL_FONT = Font(name='Arial')
WARNING_FONT = Font(size=10, bold=True, color='FF0000', name='Arial')
TITLE_FONT = Font(size=14, bold=True, color='FFFFFF', name='Arial')
SUMMARY_TITLE_FONT = Font(size=11, bold=True, color='FFFFFF', name='Arial')
SUMMARY_TEXT_FONT = Font(size=10, name='Arial')
INST_LABEL_FONT = Font(size=10, bold=True, color='4472C4', name='Arial')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)


# ── 数据读取与计算 ────────────────────────────────────

def read_institution_data(input_file):
    """读取所有机构sheet的数据（排除首尾sheet）"""
    xl = pd.ExcelFile(input_file)
    sheet_names = xl.sheet_names[1:-1]

    institution_data = {}
    for sheet in sheet_names:
        df = pd.read_excel(input_file, sheet_name=sheet, engine='openpyxl')
        df = df[df['交易日期'].astype(str).str.contains(r'\d{4}/\d{2}/\d{2}')]
        df['交易日期'] = pd.to_datetime(df['交易日期'])
        institution_data[sheet] = df

    return institution_data, sheet_names


def calculate_monthly_summary(institution_data, sheet_names, start_date, end_date):
    """计算指定时间区间的申赎汇总"""
    monthly_data = {
        sheet: df[(df['交易日期'] >= start_date) & (df['交易日期'] <= end_date)]
        for sheet, df in institution_data.items()
    }

    fund_columns = [col for col in monthly_data[sheet_names[0]].columns
                    if col not in ['交易日期', '机构类型']]

    summary_by_inst = {inst: df[fund_columns].sum() for inst, df in monthly_data.items()}

    # 构建汇总DataFrame
    summary_dict = {'基金类型': fund_columns}
    for inst in sheet_names:
        summary_dict[f'{inst}净申赎'] = [round(summary_by_inst[inst][col], 2) for col in fund_columns]
    summary_dict['合计'] = [
        round(sum(summary_by_inst[inst][col] for inst in sheet_names), 2)
        for col in fund_columns
    ]

    df_summary = pd.DataFrame(summary_dict)

    # 合计行 - 只对大类求和
    total_row = {'基金类型': '总计'}
    for col in df_summary.columns:
        if col != '基金类型':
            total_row[col] = df_summary[df_summary['基金类型'].isin(MAJOR_CATEGORIES)][col].sum()
    df_summary = pd.concat([df_summary, pd.DataFrame([total_row])], ignore_index=True)

    return df_summary, fund_columns, monthly_data


# ── Excel格式化 ───────────────────────────────────────

def format_summary_sheet(ws, df_summary, fund_columns, start_date, end_date):
    """格式化汇总报告sheet，返回数据区域结束行号"""
    num_cols = len(df_summary.columns)
    last_col_letter = get_column_letter(num_cols)
    date_text = f"{start_date.strftime('%Y年%m月%d日')}-{end_date.strftime('%Y年%m月%d日')}"

    # A1 警告
    ws['A1'] = '该报告基于日度脱敏数据生成,数据本身不代表交易信息,请知悉!禁止转发及用于商业用途,否则将追究法律责任!'
    ws['A1'].font = WARNING_FONT
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 20

    # A2 标题
    ws['A2'] = f'基金申赎报告 - {date_text}'
    ws['A2'].font = TITLE_FONT
    ws['A2'].fill = TITLE_FILL
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'].alignment = CENTER
    ws.row_dimensions[2].height = 25

    # 表头（第4行）
    for col in range(1, num_cols + 1):
        cell = ws.cell(4, col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    # 合计行
    total_row_num = len(fund_columns) + 5
    for col in range(1, num_cols + 1):
        cell = ws.cell(total_row_num, col)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL

    # 列宽
    ws.column_dimensions['A'].width = 28
    for i in range(2, num_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # 边框 + 数值格式
    for row in ws.iter_rows(min_row=4, max_row=total_row_num, min_col=1, max_col=num_cols):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.column > 1:
                cell.number_format = '#,##0.00'
                cell.alignment = RIGHT

    return total_row_num


# ── AI摘要插入 ────────────────────────────────────────

def default_summary_generator(df_summary):
    """
    占位：AI摘要生成函数接口。

    参数:
        df_summary: pd.DataFrame - 汇总表（含基金类型列、各机构净申赎列、合计列、总计行）

    返回:
        dict: {
            'overall': str,          # 整体情况（多行文本，用换行分隔）
            'institutions': {        # 分机构分析
                '保险': str,
                '券商': str,
                ...
            }
        }

    实现时替换此函数，例如:
        def my_llm_summary(df_summary):
            prompt = f"请分析以下基金申赎数据:\\n{df_summary.to_string()}"
            response = call_llm(prompt)
            return parse_response(response)
    """
    return None


def insert_summary_to_sheet(ws, start_row, num_cols, summary_data):
    """
    将AI摘要插入到Excel表格下方。

    参数:
        ws: worksheet对象
        start_row: 插入起始行（通常为数据表结束行 + 2）
        num_cols: 总列数
        summary_data: dict，格式同 default_summary_generator 返回值
    """
    if not summary_data:
        return

    last_col = get_column_letter(num_cols)
    row = start_row

    # ── 整体情况 ──
    if 'overall' in summary_data and summary_data['overall']:
        # 标题行
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws.cell(row, 1, '整体情况')
        cell.font = SUMMARY_TITLE_FONT
        cell.fill = SUMMARY_HEADER_FILL
        cell.alignment = CENTER
        ws.row_dimensions[row].height = 22
        row += 1

        # 内容行
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws.cell(row, 1, summary_data['overall'])
        cell.font = SUMMARY_TEXT_FONT
        cell.alignment = WRAP
        cell.fill = SUMMARY_BG_FILL

        # 根据文本行数设置行高
        line_count = summary_data['overall'].count('\n') + 1
        ws.row_dimensions[row].height = max(60, line_count * 18)
        row += 2

    # ── 分机构 ──
    institutions = summary_data.get('institutions', {})
    if not institutions:
        return

    # 标题行
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = ws.cell(row, 1, '分机构')
    cell.font = SUMMARY_TITLE_FONT
    cell.fill = SUMMARY_HEADER_FILL
    cell.alignment = CENTER
    ws.row_dimensions[row].height = 22
    row += 1

    # 每行放2个机构，左半区 + 右半区
    inst_list = list(institutions.items())
    mid_col = num_cols // 2  # 左半区列数
    left_last = get_column_letter(mid_col)
    right_first_idx = mid_col + 1
    right_last = last_col

    for i in range(0, len(inst_list), 2):
        # 左侧机构标签
        name_l, text_l = inst_list[i]
        ws.merge_cells(f'A{row}:{left_last}{row}')
        cell_l = ws.cell(row, 1, name_l)
        cell_l.font = INST_LABEL_FONT
        cell_l.alignment = CENTER

        # 右侧机构标签（如果有）
        if i + 1 < len(inst_list):
            name_r, text_r = inst_list[i + 1]
            ws.merge_cells(f'{get_column_letter(right_first_idx)}{row}:{right_last}{row}')
            cell_r = ws.cell(row, right_first_idx, name_r)
            cell_r.font = INST_LABEL_FONT
            cell_r.alignment = CENTER
        row += 1

        # 左侧内容
        ws.merge_cells(f'A{row}:{left_last}{row}')
        cell_lt = ws.cell(row, 1, text_l)
        cell_lt.font = SUMMARY_TEXT_FONT
        cell_lt.alignment = WRAP
        cell_lt.fill = SUMMARY_BG_FILL

        line_count_l = text_l.count('\n') + 1

        # 右侧内容
        line_count_r = 0
        if i + 1 < len(inst_list):
            ws.merge_cells(f'{get_column_letter(right_first_idx)}{row}:{right_last}{row}')
            cell_rt = ws.cell(row, right_first_idx, text_r)
            cell_rt.font = SUMMARY_TEXT_FONT
            cell_rt.alignment = WRAP
            cell_rt.fill = SUMMARY_BG_FILL
            line_count_r = text_r.count('\n') + 1

        ws.row_dimensions[row].height = max(60, max(line_count_l, line_count_r) * 18)
        row += 1


# ── 趋势图表 ─────────────────────────────────────────

def prepare_trend_data(monthly_data, inst):
    """准备趋势分析数据（累计值）"""
    df = monthly_data[inst].copy()
    available = [c for c in MAJOR_CATEGORIES if c in df.columns]
    df_trend = df[['交易日期'] + available].copy()
    df_trend.columns = ['日期'] + available
    df_trend = df_trend.sort_values('日期').reset_index(drop=True)
    df_trend['日期'] = df_trend['日期'].dt.strftime('%Y/%-m/%-d')

    for col in available:
        df_trend[col] = df_trend[col].cumsum()
    return df_trend


def create_trend_chart(ws, df_trend, inst, start_date, end_date):
    """创建趋势折线图"""
    chart = LineChart()
    chart.title = f'{inst} - {start_date.strftime("%Y/%m/%d")} 至 {end_date.strftime("%Y/%m/%d")} 大类申赎累计趋势'
    chart.style = 2
    chart.y_axis.title = '净申赎'
    chart.height = 10
    chart.width = 20

    categories = Reference(ws, min_col=1, min_row=3, max_row=len(df_trend) + 2)
    for idx in range(2, len(df_trend.columns) + 1):
        data = Reference(ws, min_col=idx, min_row=2, max_row=len(df_trend) + 2)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend.position = 'b'
    chart.y_axis.majorGridlines = None
    return chart


def add_trend_sheets(wb, monthly_data, sheet_names, start_date, end_date):
    """添加趋势分析sheet和图表"""
    for inst in sheet_names:
        df_trend = prepare_trend_data(monthly_data, inst)
        ws = wb.create_sheet(title=inst)
        num_trend_cols = len(df_trend.columns)

        # 标题
        ws.merge_cells(f'A1:{get_column_letter(num_trend_cols)}1')
        cell = ws.cell(1, 1, f'{inst} - 大类申赎累计趋势')
        cell.font = Font(size=12, bold=True, color='FFFFFF', name='Arial')
        cell.fill = TITLE_FILL
        cell.alignment = CENTER

        # 表头
        for c_idx, col_name in enumerate(df_trend.columns, 1):
            cell = ws.cell(2, c_idx, col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER

        # 数据
        for r_idx, row in enumerate(df_trend.itertuples(index=False), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, value)
                if c_idx > 1:
                    cell.number_format = '#,##0.00'
                    cell.alignment = RIGHT

        # 列宽
        ws.column_dimensions['A'].width = 12
        for i in range(2, num_trend_cols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        # 图表
        chart = create_trend_chart(ws, df_trend, inst, start_date, end_date)
        ws.add_chart(chart, f'A{len(df_trend) + 5}')


# ── 主函数 ────────────────────────────────────────────

def generate_monthly_report(
    input_file,
    output_file,
    start_date,
    end_date,
    add_trend_charts=True,
    summary_generator=None,
):
    """
    生成基金申赎月度报告。

    参数:
        input_file: 原始数据Excel文件
        output_file: 输出报告文件路径
        start_date, end_date: 日期区间 (str 或 datetime)
        add_trend_charts: 是否添加趋势图sheet
        summary_generator: 可选，AI摘要生成函数，签名 f(df_summary) -> dict
    """
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)

    # 1. 读取数据
    print('正在读取数据...')
    institution_data, sheet_names = read_institution_data(input_file)

    # 2. 计算汇总
    print(f'正在计算{start_date.date()}至{end_date.date()}汇总...')
    df_summary, fund_columns, monthly_data = calculate_monthly_summary(
        institution_data, sheet_names, start_date, end_date
    )

    # 3. 写入Excel
    print('正在生成报告...')
    sheet_name = f'{start_date.strftime("%Y%m%d")}-{end_date.strftime("%Y%m%d")}申赎汇总'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)

    # 4. 格式化
    wb = load_workbook(output_file)
    ws = wb[sheet_name]
    total_row_num = format_summary_sheet(ws, df_summary, fund_columns, start_date, end_date)

    # 5. 插入AI摘要
    gen_fn = summary_generator or default_summary_generator
    summary_data = gen_fn(df_summary)
    if summary_data:
        print('正在插入摘要...')
        insert_summary_to_sheet(ws, total_row_num + 2, len(df_summary.columns), summary_data)

    # 6. 趋势图表
    if add_trend_charts:
        print('正在生成趋势图...')
        add_trend_sheets(wb, monthly_data, sheet_names, start_date, end_date)

    wb.save(output_file)

    print(f'\n✓ 报告生成完成: {output_file}')
    print(f'  机构: {len(sheet_names)}个')
    for inst in sheet_names:
        print(f'    {inst}: {len(monthly_data[inst])}个交易日')
    print(f'  基金类型: {len(fund_columns)}个')
    if add_trend_charts:
        print(f'  趋势图: {len(sheet_names)}个sheet')

    return output_file
